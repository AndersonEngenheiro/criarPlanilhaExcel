from openpyxl import Workbook


arquivo = Workbook()

plan0 = arquivo.active

plan_final = arquivo.create_sheet('final') # a última por padrão
plan_primeira = arquivo.create_sheet('primeira', 0) # primeira posição
plan_penultima = arquivo.create_sheet('penúltima', -1) # penúltima position

plan0.title = 'MeuArquivo'

plan0.sheet_properties.tabColor = '1079BA'

print(arquivo.sheetnames)

alvo = arquivo.copy_worksheet(plan0)

print(arquivo.sheetnames)



plan0['A1'] = 5
plan0['A2'] = 7
plan0['A3'] = 0
plan0['A4'] = 1
plan0['A5'] = 6
plan_final['B5'] = 55
plan_final['B7'] = 88




# for row in plan0.values:
#     for value in row:
#         print(value)

# forma 2

for row in plan0.iter_rows(min_row=2, max_col=6, max_row=4, values_only=True):
    print(row)

c = plan_primeira['D7']
c.value = 'anderson'

arquivo.save('info.xlsx')